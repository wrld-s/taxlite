"""Excel output generation matching the reimbursement template."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from .scanner import ReceiptData


HEADERS = [
    "DATE",
    "PAYEE/ VENDOR NAME",
    "TIN #",
    "CASH AMOUNT",
    "12% VAT",
    "NET",
    "ADDRESS",
    "Reference #/ OR #/ INV #",
    "PARTICULARS",
]

COL_WIDTHS = [14, 40, 22, 16, 14, 16, 30, 24, 18]


def generate_excel(
    receipts: list[ReceiptData],
    output_path: Path,
    month: Optional[datetime] = None,
):
    """Generate an Excel file matching the reimbursement template format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "INPUT"

    # Styles
    title_font = Font(name="Calibri", size=14, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    date_format = "YYYY-MM-DD"
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Column widths
    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Row 1: Title
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "SUMMARY OF EXPENSES"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    # Row 7: Month
    ws["A7"] = "MONTH OF"
    ws["A7"].font = Font(bold=True)
    if month:
        ws["B7"] = month
        ws["B7"].number_format = "MMMM YYYY"
    else:
        # Auto-detect from receipts
        dates = [r.date for r in receipts if r.date]
        if dates:
            try:
                first_date = datetime.strptime(sorted(dates)[0], "%Y-%m-%d")
                ws["B7"] = first_date.replace(day=1)
                ws["B7"].number_format = "MMMM YYYY"
            except ValueError:
                pass

    # Row 8: Subtotal formulas
    data_start = 10
    data_end = data_start + len(receipts) - 1 if receipts else data_start
    ws["C8"] = "SUBTOTAL"
    ws["C8"].font = Font(bold=True)
    ws["D8"] = f"=SUM(D{data_start}:D{data_end})"
    ws["D8"].number_format = money_format
    ws["D8"].font = Font(bold=True)
    ws["E8"] = f"=SUM(E{data_start}:E{data_end})"
    ws["E8"].number_format = money_format
    ws["E8"].font = Font(bold=True)
    ws["F8"] = f"=SUM(F{data_start}:F{data_end})"
    ws["F8"].number_format = money_format
    ws["F8"].font = Font(bold=True)

    # Row 9: Headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (starting at row 10)
    # Sort receipts by date
    def sort_key(r):
        if r.date:
            try:
                return datetime.strptime(r.date, "%Y-%m-%d")
            except ValueError:
                pass
        return datetime.max

    sorted_receipts = sorted(receipts, key=sort_key)

    for i, receipt in enumerate(sorted_receipts):
        row = data_start + i

        # A: Date
        if receipt.date:
            try:
                dt = datetime.strptime(receipt.date, "%Y-%m-%d")
                ws.cell(row=row, column=1, value=dt).number_format = date_format
            except ValueError:
                ws.cell(row=row, column=1, value=receipt.date)
        else:
            ws.cell(row=row, column=1, value="")

        # B: Vendor name
        ws.cell(row=row, column=2, value=receipt.vendor_name)

        # C: TIN
        ws.cell(row=row, column=3, value=receipt.tin)

        # D: Cash amount
        ws.cell(row=row, column=4, value=receipt.total_amount).number_format = money_format

        # E: 12% VAT
        ws.cell(row=row, column=5, value=receipt.vat_amount).number_format = money_format

        # F: NET (VAT / 0.12, or same as total if VAT is 0)
        if receipt.vat_amount > 0:
            ws.cell(row=row, column=6).value = f"=E{row}/0.12"
        else:
            ws.cell(row=row, column=6, value=receipt.total_amount)
        ws.cell(row=row, column=6).number_format = money_format

        # G: Address
        ws.cell(row=row, column=7, value=receipt.address)

        # H: Reference number
        ws.cell(row=row, column=8, value=receipt.receipt_number)

        # I: Particulars
        ws.cell(row=row, column=9, value=receipt.items_description)

        # Apply border to all data cells
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border

    wb.save(str(output_path))
    return output_path
